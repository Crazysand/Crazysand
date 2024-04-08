import crazysand.common as cc
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
import os


# 定义一个类，用于处理Excel中的图片相关操作
class MxExImage:

    def __init__(self, image_path):
        # 初始化时存储图片路径
        self.image_path = image_path


# 定义一个类，用于处理Excel文件的操作
class MxExcel:

    def __init__(self, excel_name, load=True, sheet_name='Sheet'):
        # 初始化时指定Excel文件名，是否加载现有文件，以及默认的工作表名称
        self.excel_name = excel_name
        # 检查文件是否存在，若存在且选择加载，则加载工作簿
        if os.path.exists(excel_name) and load:
            self.workbook = load_workbook(excel_name)
        else:
            # 否则创建一个新的工作簿，该工作簿默认包含一个名为'Sheet'的工作表
            self.workbook = Workbook()
        # 如果指定的工作表名称不是默认的'Sheet'，则创建一个新的工作表
        if sheet_name != 'Sheet':
            self.sheet = self.workbook.create_sheet(title=sheet_name)
        else:
            # 否则，使用默认的工作表
            self.sheet = self.workbook[sheet_name]

    def append(self, row_data: list):
        # 添加一行数据，如果数据中包含图片，则进行特殊处理
        insert_row = self.last_row_with_content + 1
        for index, value in enumerate(row_data):
            if isinstance(value, MxExImage):
                # 如果值是MxExImage实例，则在指定位置插入图片，并将单元格内容置空
                row_data[index] = ''
                self.insert_image(self.sheet, insert_row, index + 1, value.image_path)
            else:
                # 否则，直接在单元格中写入值
                self.sheet.cell(row=insert_row, column=index + 1, value=value)
        return self

    def append_by_headers(self, data: dict):
        # 根据表头添加一行数据，如果数据中包含图片，则进行特殊处理
        insert_row = self.last_row_with_content + 1
        row_data = [data.get(header) for header in self.headers]
        for index, value in enumerate(row_data):
            if isinstance(value, MxExImage):
                # 如果值是MxExImage实例，则在指定位置插入图片，并将单元格内容置空
                row_data[index] = ''
                self.insert_image(self.sheet, insert_row, index + 1, value.image_path)
            else:
                # 否则，直接在单元格中写入值
                self.sheet.cell(row=insert_row, column=index + 1, value=value)
        return self

    def save(self):
        # 保存工作簿到文件
        self.workbook.save(self.excel_name)

    def set_cells_size(self, end_row, end_column, cell_width_px, cell_height_px, start_row=1,
                       start_column=1):
        """
        以像素为单位调整Excel表格中指定范围内单元格的大小。

        参数:
        - workbook_path: str, Excel工作簿的路径。
        - sheet_name: str, 工作表名称。
        - start_row: int, 起始行号。
        - end_row: int, 结束行号。
        - start_column: int, 起始列号（以1为起始，而非A、B这样的列标识）。
        - end_column: int, 结束列号。
        - cell_height_px: int, 单元格的高度（像素）。
        - cell_width_px: int, 单元格的宽度（像素）。
        """
        # 转换像素到Excel单位
        cell_height = cell_height_px / 1.3333  # 从像素转换到磅
        cell_width = cell_width_px / 7  # 这是一个近似值，可能需要根据字体调整

        # 设置指定区域内的单元格尺寸
        for row in range(start_row, end_row + 1):
            self.sheet.row_dimensions[row].height = cell_height
        for col in range(start_column, end_column + 1):
            col_letter = self.sheet.cell(row=start_row, column=col).column_letter
            self.sheet.column_dimensions[col_letter].width = cell_width

    @property
    def headers(self):
        return [cell.value for cell in self.sheet[1]]

    @property
    def last_row_with_content(self):
        last_row_with_data = 0
        for row in self.sheet.iter_rows():
            if any(cell.value is not None for cell in row):
                last_row_with_data = row[0].row

        return last_row_with_data

    @staticmethod
    def insert_image(sheet, row, column, image_path):
        """
        将图片嵌入到Excel工作表的指定位置。

        参数:
        - sheet: 工作表对象，图片将被添加到这个工作表。
        - row: int, 图片要插入的行号。
        - column: int, 图片要插入的列号。
        - image_path: str, 要插入的图片的文件路径。
        """
        # 获取列的字母表示
        column_letter = sheet.cell(row=row, column=column).column_letter

        # 创建图片对象
        img = Image(image_path)

        # 设置图片的锚点为指定的单元格位置，例如 "A1"
        anchor = f"{column_letter}{row}"
        img.anchor = anchor

        # 将图片添加到工作表
        sheet.add_image(img)


if __name__ == '__main__':
    lis = [
        (1, 93.59, '清华大学', 'public', '教育部', ['综合'],
         ['985', '211', '双一流', '国重点', '保研', '研究生院', 'C9', '建筑 老八校'], '北京', '海淀区',
         'http://img1.youzy.cn/content/media/thumbs/p00189999.jpeg'),
        (1, 93.59, '清华大学医学部', 'public', '教育部', ['医药'], ['985', '211', '双一流', '国重点', '保研'], '北京',
         '东城区', 'http://img1.youzy.cn/content/media/thumbs/p00200807.jpeg'),
        (1, 93.59, '清华大学美术学院', 'public', '教育部', ['艺术'], ['985', '211', '双一流', '国重点', '保研'], '北京',
         '海淀区', 'http://img1.youzy.cn/content/media/thumbs/p00181049.png'),
        (2, 92.13, '北京大学', 'public', '教育部', ['综合'],
         ['985', '211', '双一流', '国重点', '保研', '研究生院', 'C9', '五院 四系'], '北京', '海淀区',
         'http://img1.youzy.cn/content/media/thumbs/p00189371.jpeg')
    ]

    mxe = MxExcel('demo.xlsx', load=False)
    mxe.set_cells_size(end_row=100, end_column=5, cell_width_px=100, cell_height_px=100)

    mxe.append(['排名', '分数', '名称', 'logo'])

    for _ in lis:
        url = _[-1]
        path = cc.SourceRequest.image(url, suffix='.png', parent_path='./images/')
        cc.resize_image(path, target_width=100)
        mxe.append_by_headers({
            '排名': _[0],
            '分数': _[1],
            '名称': _[2],
            'logo': MxExImage(path)
        })
    mxe.save()
